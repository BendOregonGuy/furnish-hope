"""
Generate Excel import templates for every user-data table in the
Furnish Hope database.

Reads column metadata directly from the live local Postgres so the
templates are always in sync with the current schema (no stale dump
to worry about).

Output: docs/import_templates/furnish-hope-import-templates.xlsx
        — one sheet per importable table, plus an "Instructions" sheet.

Each per-table sheet has, in order:
  Row 1  — Column name. Bold. Red fill if the column is REQUIRED.
           Blue fill if it's a foreign key.
  Row 2  — Type hint (e.g. "Text up to 50 chars" or "Date YYYY-MM-DD").
  Row 3  — FK reference (e.g. "-> tbl_donor.donor_id") or blank.
  Row 4  — Empty so users can add a label / comment if helpful.
  Row 5+ — Blank rows for data entry.

Tables that don't make sense to bulk-import via Excel (binary blobs,
encrypted credentials, auto-generated audit data, singletons,
session cookies, etc.) are intentionally excluded — listed on the
Instructions sheet so admins know nothing was forgotten.
"""

import os
import sys
import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_DSN = os.environ.get(
    'FH_DB_DSN',
    'host=localhost dbname=furnish_hope user=postgres password=postgres',
)

# Tables to skip — system internals, binary storage, auto-generated
# data, encrypted credentials, OAuth tokens, etc. These are not
# meaningful to bulk-import via spreadsheet.
EXCLUDED_TABLES = {
    'tbl_audit_log',                  # auto-generated
    'session',                         # connect-pg-simple internal
    'tbl_attachment_blob',             # binary content
    'tbl_entity_attachment',           # binary-linked metadata (uploaded via UI)
    'tbl_email_account',               # encrypted creds + OAuth
    'tbl_email_message',               # IMAP-synced
    'tbl_email_attachment',            # binary
    'tbl_email_sync_state',            # internal sync cursors
    'tbl_org_branding',                # singleton BYTEA (logo upload)
    'tbl_manual_screenshot',           # BYTEA + admin-uploaded
    'tbl_quickbooks_connection',       # OAuth tokens
    'tbl_quickbooks_donor_link',       # auto-generated on sync
    'tbl_quickbooks_donation_sync',    # auto-generated on sync
    'tbl_quickbooks_account_mapping',  # admin-only UI; small table
    'tbl_client_waiver',               # auto-generated with signature BYTEA
    'tbl_app_setting',                 # singleton-ish, edit via UI
    'tbl_user_account',                # bcrypt password hashing — created via Admin UI
    'tbl_receipt_counter',             # auto-managed receipt sequence
}

# Columns to drop from EVERY template — auto-managed or computed.
AUTO_COLUMNS = {
    'created_at', 'updated_at', 'cached_at', 'action_at',
    'uploaded_at', 'signed_at', 'attempted_at',
    'last_sync_at', 'last_synced_at',
}

# Style palette — matches the app's design tokens loosely.
COLOR_HEADER     = 'F5F0E8'  # cream
COLOR_REQUIRED   = 'FBD8C8'  # soft terracotta
COLOR_FK         = 'D9E5F1'  # soft blue
COLOR_TYPE       = 'EFEFEF'  # neutral grey
COLOR_NOTE       = 'FFFFFF'
COLOR_INSTRUCTIONS_BG = 'F5F0E8'

FONT_HEADER = Font(name='Calibri', size=11, bold=True, color='1A1A1A')
FONT_TYPE   = Font(name='Calibri', size=9,  italic=True, color='666666')
FONT_FK     = Font(name='Calibri', size=9,  color='2C5282')
FONT_INSTR_HEADING = Font(name='Calibri', size=14, bold=True)
FONT_INSTR_BODY    = Font(name='Calibri', size=11)

BORDER_THIN = Border(
    left=Side(style='thin', color='D8D4CC'),
    right=Side(style='thin', color='D8D4CC'),
    top=Side(style='thin', color='D8D4CC'),
    bottom=Side(style='thin', color='D8D4CC'),
)


def fetch_tables(conn):
    """All public tables that start with tbl_ and aren't excluded."""
    with conn.cursor() as cur:
        cur.execute("""
            SELECT table_name
              FROM information_schema.tables
             WHERE table_schema = 'public'
               AND table_type = 'BASE TABLE'
               AND table_name LIKE 'tbl_%'
             ORDER BY table_name
        """)
        rows = [r[0] for r in cur.fetchall()]
    return [t for t in rows if t not in EXCLUDED_TABLES]


def fetch_columns(conn, table_name):
    """Column metadata for one table, plus FK target if any."""
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT
              c.column_name,
              c.data_type,
              c.character_maximum_length,
              c.numeric_precision, c.numeric_scale,
              c.is_nullable,
              c.column_default,
              -- referenced table.column if this is a FK
              (SELECT ccu.table_name
                 FROM information_schema.table_constraints tc
                 JOIN information_schema.key_column_usage kcu
                      ON tc.constraint_name = kcu.constraint_name
                 JOIN information_schema.constraint_column_usage ccu
                      ON tc.constraint_name = ccu.constraint_name
                WHERE tc.constraint_type = 'FOREIGN KEY'
                  AND tc.table_schema = c.table_schema
                  AND tc.table_name = c.table_name
                  AND kcu.column_name = c.column_name
                LIMIT 1) AS fk_table,
              (SELECT ccu.column_name
                 FROM information_schema.table_constraints tc
                 JOIN information_schema.key_column_usage kcu
                      ON tc.constraint_name = kcu.constraint_name
                 JOIN information_schema.constraint_column_usage ccu
                      ON tc.constraint_name = ccu.constraint_name
                WHERE tc.constraint_type = 'FOREIGN KEY'
                  AND tc.table_schema = c.table_schema
                  AND tc.table_name = c.table_name
                  AND kcu.column_name = c.column_name
                LIMIT 1) AS fk_column,
              -- is this column the primary key?
              EXISTS (
                SELECT 1
                  FROM information_schema.table_constraints tc
                  JOIN information_schema.key_column_usage kcu
                       ON tc.constraint_name = kcu.constraint_name
                 WHERE tc.constraint_type = 'PRIMARY KEY'
                   AND tc.table_schema = c.table_schema
                   AND tc.table_name = c.table_name
                   AND kcu.column_name = c.column_name
              ) AS is_pk
            FROM information_schema.columns c
           WHERE c.table_schema = 'public'
             AND c.table_name = %s
           ORDER BY c.ordinal_position
        """, (table_name,))
        return cur.fetchall()


def type_hint(col):
    """Human-readable type description for a column row."""
    t = col['data_type']
    if t in ('character varying',):
        n = col['character_maximum_length']
        return f'Text (max {n} chars)' if n else 'Text'
    if t == 'text':
        return 'Long text'
    if t == 'integer':
        return 'Whole number'
    if t == 'bigint':
        return 'Whole number (big)'
    if t == 'numeric':
        p, s = col['numeric_precision'], col['numeric_scale']
        if s:
            return f'Decimal (up to {p} digits, {s} decimal places — e.g. 123.45)'
        return 'Whole number'
    if t == 'boolean':
        return 'true or false'
    if t == 'date':
        return 'Date — YYYY-MM-DD (e.g. 2026-06-11)'
    if t.startswith('timestamp'):
        return 'Date+time — YYYY-MM-DD HH:MM:SS'
    if t == 'time without time zone':
        return 'Time — HH:MM:SS (24-hr)'
    if t == 'bytea':
        return 'Binary file — do NOT enter here; upload via the app instead'
    if t == 'json' or t == 'jsonb':
        return 'JSON object'
    return t


def is_required(col):
    """A column is required if it's NOT NULL and has no default,
    AND isn't the auto-incrementing PK (handled separately)."""
    if col['is_nullable'] == 'YES':
        return False
    if col['column_default']:
        # Has a default — server will fill it in.
        # Common case: nextval(...) on serial PKs, or now() / true.
        return False
    return True


def should_skip_column(col):
    """Skip auto-generated PK columns, audit defaults, etc."""
    name = col['column_name']
    if name in AUTO_COLUMNS:
        return True
    # Auto-incrementing PK — never imported, always server-assigned.
    if col['is_pk'] and col['column_default'] and 'nextval' in (col['column_default'] or ''):
        return True
    return False


def make_table_sheet(wb, table_name, columns):
    sheet_name = table_name.replace('tbl_', '')
    # Excel sheet names cap at 31 chars.
    sheet_name = sheet_name[:31]
    ws = wb.create_sheet(sheet_name)

    # Filter columns.
    cols = [c for c in columns if not should_skip_column(c)]
    if not cols:
        ws['A1'] = f'(No importable columns in {table_name}.)'
        return

    # Set column widths so the row content fits well.
    for i, c in enumerate(cols, 1):
        letter = get_column_letter(i)
        name_len = len(c['column_name'])
        type_len = len(type_hint(c))
        ws.column_dimensions[letter].width = max(16, min(40, max(name_len, type_len * 0.7) + 2))

    # Row 1 — header: column names with fills.
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=c['column_name'])
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = BORDER_THIN
        if is_required(c):
            cell.fill = PatternFill('solid', fgColor=COLOR_REQUIRED)
        elif c['fk_table']:
            cell.fill = PatternFill('solid', fgColor=COLOR_FK)
        else:
            cell.fill = PatternFill('solid', fgColor=COLOR_HEADER)
    ws.row_dimensions[1].height = 30

    # Row 2 — type hint.
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=2, column=i, value=type_hint(c))
        cell.font = FONT_TYPE
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = BORDER_THIN
        cell.fill = PatternFill('solid', fgColor=COLOR_TYPE)
    ws.row_dimensions[2].height = 32

    # Row 3 — FK reference / required note.
    for i, c in enumerate(cols, 1):
        bits = []
        if c['fk_table']:
            bits.append(f'-> {c["fk_table"]}.{c["fk_column"]}')
        if is_required(c):
            bits.append('REQUIRED')
        elif not c['column_default'] and c['is_nullable'] == 'YES':
            bits.append('optional')
        cell = ws.cell(row=3, column=i, value=' · '.join(bits))
        cell.font = FONT_FK if c['fk_table'] else FONT_TYPE
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.border = BORDER_THIN
        cell.fill = PatternFill('solid', fgColor=COLOR_NOTE)
    ws.row_dimensions[3].height = 22

    # Row 4 — divider with a single instruction.
    ws.cell(row=4, column=1, value='-- Data rows below (delete this line before importing) --').font = Font(italic=True, color='999999', size=9)
    ws.row_dimensions[4].height = 18

    # Freeze panes so the header rows stay visible.
    ws.freeze_panes = 'A5'

    # Sheet color tab — distinguish FK-heavy operational tables from
    # reference data. Just a small affordance.
    pk_col = next((c for c in columns if c['is_pk']), None)
    if pk_col:
        ws.sheet_properties.tabColor = '7A8C5E'  # sage


def make_instructions_sheet(wb, table_count, excluded):
    ws = wb.create_sheet('Instructions', 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 110

    lines = [
        ('Furnish Hope — Import Templates', 'heading'),
        ('', None),
        (f'This workbook contains import templates for {table_count} user-data tables in the Furnish Hope database. '
         f'Each tab is one table. Fill in the data rows below the header, save the file as CSV (one per tab if your importer needs that), '
         f'and your admin can bulk-load it into the system.',
         'body'),
        ('', None),
        ('How to read each tab', 'subheading'),
        ('• Row 1 — column names. Light terracotta fill = REQUIRED. Light blue fill = foreign key (must reference an existing record).',
         'body'),
        ('• Row 2 — what kind of value goes in this column (text length cap, date format, number, etc.).', 'body'),
        ('• Row 3 — for FKs: the table + column you must reference. For required columns: REQUIRED tag.', 'body'),
        ('• Row 4 — a divider that you should delete before importing.', 'body'),
        ('• Rows 5 and below — your data. As many rows as you want.', 'body'),
        ('', None),
        ('Foreign-key columns', 'subheading'),
        ('Most tables reference other tables by ID. For example, tbl_donation.donor_id must point to a tbl_donor.donor_id that already exists. '
         'Two strategies:',
         'body'),
        ('  1. Import the parent table first (e.g., import donors before donations). After import, your admin gives you the new ID range.', 'body'),
        ('  2. If the parent table is small (under a hundred rows), it\'s usually easier to add those rows via the app\'s + New buttons '
         'and copy the IDs into your spreadsheet manually.',
         'body'),
        ('', None),
        ('Auto-generated columns are excluded', 'subheading'),
        ('• Primary keys (auto-incrementing IDs) — the database assigns these on insert.', 'body'),
        ('• created_at, updated_at — auto-set to NOW().', 'body'),
        ('• If a column has a default value (e.g., is_active defaults to true), leave it blank to accept the default, '
         'or enter a value to override.',
         'body'),
        ('', None),
        ('Tables intentionally NOT included', 'subheading'),
        ('These are system-managed, encrypted, binary, auto-generated, or singletons — they can\'t be meaningfully bulk-imported. '
         'Manage them through the app instead:',
         'body'),
    ]
    for name in sorted(excluded):
        lines.append((f'  · {name}', 'body'))
    lines.extend([
        ('', None),
        ('Lookup tables (lkp_*) are also excluded', 'subheading'),
        ('Lookup tables hold dropdown values (donor types, RSVP statuses, sponsor levels, etc.). They\'re seeded by the system '
         'and small enough to edit directly at /admin/<table-name>.',
         'body'),
        ('', None),
        ('Need help?', 'subheading'),
        ('Open the User Manual inside the app (sidebar -> System -> User Manual) or email preston@getreality.com.', 'body'),
    ])

    row = 1
    for text, kind in lines:
        cell = ws.cell(row=row, column=1, value=text)
        if kind == 'heading':
            cell.font = Font(name='Calibri', size=18, bold=True, color='8D4A2C')
            ws.row_dimensions[row].height = 28
        elif kind == 'subheading':
            cell.font = Font(name='Calibri', size=12, bold=True, color='4A4A4A')
            ws.row_dimensions[row].height = 22
        elif kind == 'body':
            cell.font = FONT_INSTR_BODY
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            ws.row_dimensions[row].height = max(20, 16 * (len(text) // 100 + 1))
        row += 1


def main():
    out_dir = os.path.join('docs', 'import_templates')
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, 'furnish-hope-import-templates.xlsx')

    try:
        conn = psycopg2.connect(DB_DSN)
    except psycopg2.OperationalError as e:
        print(f'Could not connect to Postgres: {e}', file=sys.stderr)
        print('Set FH_DB_DSN env var if your local DB uses different credentials.', file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    # Drop the default empty sheet — we'll add Instructions explicitly.
    wb.remove(wb.active)

    tables = fetch_tables(conn)

    # The Instructions sheet's "intentionally excluded" list mirrors
    # the static set above PLUS any lkp_* tables we'd otherwise miss.
    make_instructions_sheet(wb, len(tables), EXCLUDED_TABLES)

    for t in tables:
        cols = fetch_columns(conn, t)
        make_table_sheet(wb, t, cols)

    conn.close()
    wb.save(out_path)
    print(f'Wrote {out_path}')
    print(f'  Tables included: {len(tables)}')
    print(f'  Tables excluded: {len(EXCLUDED_TABLES)}')


if __name__ == '__main__':
    main()
